from __future__ import annotations

import asyncio
from datetime import datetime
from io import BytesIO
from collections import defaultdict
from dataclasses import dataclass, field

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import BotCommand, BufferedInputFile, CallbackQuery, InputMediaPhoto, Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bot.config import load_config
from bot.dates import parse_return_plan
from bot.keyboards import condition_menu, model_by_index, model_category, model_legend, model_menu, skip_take_comment_menu, start_menu
from bot.parser import parse_approval, parse_return, parse_take
from bot.rules import check_take_rules, validate_approval
from bot.storage import Storage


HELP_TEXT = """Форматы:

Взял JAC J7 Н537РА126 на 2 дня, возврат 29 августа
Взял JAC J7 Н537РА126 на 1 день, ночная смена, возврат завтра

Возврат:
Вернул JAC J7 Н537РА126, аренда 2 дня

К выдаче и возврату нужно приложить 5 фото: 4 стороны авто и приборная панель с топливом.
"""


class TakeFlow(StatesGroup):
    model = State()
    plate = State()
    return_text = State()
    comment = State()
    photos = State()


class ProfileFlow(StatesGroup):
    name = State()


class ReturnFlow(StatesGroup):
    plate = State()
    condition = State()
    photos = State()


class TestDriveFlow(StatesGroup):
    model = State()
    plate = State()
    comment = State()
    photos = State()


@dataclass
class MediaGroupBuffer:
    messages: list[Message] = field(default_factory=list)
    task: asyncio.Task | None = None


config = load_config()
storage = Storage(config.db_path)
dp = Dispatcher(storage=MemoryStorage())
media_groups: dict[str, MediaGroupBuffer] = defaultdict(MediaGroupBuffer)


def username(message: Message) -> str | None:
    return message.from_user.username.lower() if message.from_user and message.from_user.username else None


def employee_name(message: Message) -> str:
    if not message.from_user:
        return "Неизвестный сотрудник"
    saved_name = storage.get_employee_name(message.from_user.id)
    if saved_name:
        return saved_name
    return message.from_user.full_name


def has_employee_name(user) -> bool:
    saved_name = storage.get_employee_name(user.id)
    if saved_name:
        return True
    full_name = " ".join((user.full_name or "").split())
    if not full_name or full_name.lower() == (user.username or "").lower():
        return False
    return len(full_name.split()) >= 2


def is_approver(message: Message) -> bool:
    return (username(message) or "") in config.approver_usernames or (username(message) or "") in config.admin_usernames


def is_report_user(message: Message) -> bool:
    current_username = username(message) or ""
    return (
        current_username in config.report_usernames
        or current_username in config.admin_usernames
        or current_username in config.approver_usernames
    )


def is_report_callback(callback: CallbackQuery) -> bool:
    current_username = callback.from_user.username.lower() if callback.from_user and callback.from_user.username else ""
    return (
        current_username in config.report_usernames
        or current_username in config.admin_usernames
        or current_username in config.approver_usernames
    )


def photo_count(messages: list[Message]) -> int:
    return sum(1 for message in messages if message.photo)


def photo_file_ids(messages: list[Message]) -> list[str]:
    return [message.photo[-1].file_id for message in messages if message.photo]


async def send_to_work_chat(bot: Bot, file_ids: list[str], summary: str) -> None:
    work_chat_id = effective_work_chat_id()
    if work_chat_id is None:
        return
    if not file_ids:
        await bot.send_message(chat_id=work_chat_id, text=summary)
        return
    media = [
        InputMediaPhoto(media=file_id, caption=summary if index == 0 else None)
        for index, file_id in enumerate(file_ids)
    ]
    await bot.send_media_group(chat_id=work_chat_id, media=media)


def effective_work_chat_id() -> int | None:
    if config.work_chat_id is not None:
        return config.work_chat_id
    value = storage.get_setting("work_chat_id")
    return int(value) if value else None


def rental_price(model: str, days: int, planned_return_at: str | None, started_at: str | None = None) -> tuple[int, int]:
    category = model_category(model)
    day_rate = 1200 if category == "ЭКОНОМ" else 1700
    half_day_rate = 600 if category == "ЭКОНОМ" else 900
    normalized_days = max(1, int(days or 1))

    if planned_return_at:
        try:
            start = datetime.fromisoformat(started_at) if started_at else datetime.now()
            planned = datetime.fromisoformat(planned_return_at)
            hours = (planned - start).total_seconds() / 3600
            if 10 <= hours <= 14:
                return half_day_rate, half_day_rate
        except ValueError:
            pass

    return day_rate, day_rate * normalized_days


def format_dt(value: str | None) -> str:
    if not value:
        return "не указана"
    try:
        return datetime.fromisoformat(value).strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return value


def user_label(full_name: str, user_name: str | None, user_id: int) -> str:
    nick = f"@{user_name}" if user_name else f"id {user_id}"
    return f"{full_name} ({nick})"


def comment_text(comment: str | None) -> str:
    return comment.strip() if comment and comment.strip() else "не указано"


def report_comment(take_comment: str | None, return_comment: str | None) -> str:
    parts = []
    if take_comment and take_comment.strip():
        parts.append(f"Взятие: {take_comment.strip()}")
    if return_comment and return_comment.strip():
        parts.append(f"Сдача: {return_comment.strip()}")
    return "\n".join(parts)


def report_days(days: int) -> str | int:
    return "тест драйв" if int(days or 0) == 0 else days


def format_application(
    title: str,
    full_name: str,
    user_name: str | None,
    user_id: int,
    model: str,
    plate: str,
    created_at: str | None,
    planned_return_at: str | None,
    return_text: str | None,
    days: int,
    comment: str | None,
    total: int | None,
) -> str:
    planned = format_dt(planned_return_at) if planned_return_at else (return_text or "не указана")
    amount = f"{total} руб." if total is not None else "не рассчитана"
    return "\n".join(
        [
            title,
            f"ФИО: {user_label(full_name, user_name, user_id)}",
            f"Гос номер авто: {plate}",
            f"Марка модель: {model}",
            f"Дата взятия: {format_dt(created_at)}",
            f"Планируемая дата возврата: {planned}",
            f"Кол-во дней: {report_days(days)}",
            f"Комментарий: {comment_text(comment)}",
            f"Сумма аренды: {amount}",
        ]
    )


async def process_messages(messages: list[Message]) -> None:
    text_message = next((message for message in messages if message.caption or message.text), messages[0])
    text = text_message.caption or text_message.text or ""
    photos = photo_count(messages)

    take = parse_take(text)
    if take:
        if photos < 5:
            await text_message.reply(
                f"Для выдачи нужно 5 фото: 4 стороны авто и приборная панель с топливом. Сейчас вижу: {photos}. "
                "Заявку не фиксирую, пришлите полный комплект через /car."
            )
            return
        check = check_take_rules(config, take, username(text_message))
        rate, total = rental_price(take.model, take.days, None)
        rental_id = storage.create_take(
            chat_id=text_message.chat.id,
            message_id=text_message.message_id,
            user_id=text_message.from_user.id,
            username=username(text_message),
            employee_name=employee_name(text_message),
            model=take.model,
            plate=take.plate,
            days=take.days,
            return_text=take.return_text,
            planned_return_at=None,
            night_shift=take.is_night_shift,
            photo_count=photos,
            take_comment=None,
            rate=rate,
            total=total,
        )
        rental = storage.get_by_id(rental_id)
        public_no = rental.rental_no if rental and rental.rental_no else rental_id
        notes = [f"Зафиксировал выдачу #{public_no}: {take.model} {take.plate}, {take.days} дн., сумма {total} руб."]
        if not check.allowed:
            notes.append("По правилам выдача не проходит: " + " ".join(check.warnings))
        elif check.warnings:
            notes.extend(check.warnings)
        notes.append("Фото отправлены в рабочий чат." if effective_work_chat_id() else "Рабочий чат не задан, фото никуда не отправлены.")
        await send_to_work_chat(
            text_message.bot,
            photo_file_ids(messages),
            format_application(
                f"Выдача #{public_no}",
                employee_name(text_message),
                username(text_message),
                text_message.from_user.id,
                take.model,
                take.plate,
                rental.created_at if rental else None,
                rental.planned_return_at if rental else None,
                take.return_text,
                take.days,
                None,
                total,
            ),
        )
        summary = await text_message.reply("\n".join(notes))
        storage.update_take_message_id(rental_id, summary.message_id)
        return

    returned = parse_return(text)
    if returned:
        if photos < 5:
            await text_message.reply(
                f"Для возврата нужно 5 фото: 4 стороны авто и приборная панель с топливом. Сейчас вижу: {photos}. "
                "Возврат не фиксирую, пришлите полный комплект через /car."
            )
            return
        if not returned.comment:
            await text_message.reply(
                "Для возврата нужен комментарий о недостатках или отметка `Машина полностью исправна` в сценарии /car. "
                "Возврат не фиксирую."
            )
            return
        rental = storage.mark_returned(
            text_message.chat.id,
            text_message.message_id,
            returned.plate,
            photos,
            return_comment=returned.comment or None,
            condition_status="comment" if returned.comment else None,
        )
        if rental is None:
            await text_message.reply("Не нашел активную выдачу по этому номеру. Проверьте номер или формат.")
            return
        model = returned.model or rental.model
        days = returned.days or rental.days
        notes = [f"Зафиксировал возврат: {model} {returned.plate}, аренда {days} дн., сумма {rental.total or 0} руб."]
        notes.append("Фото отправлены в рабочий чат." if effective_work_chat_id() else "Рабочий чат не задан, фото никуда не отправлены.")
        await send_to_work_chat(
            text_message.bot,
            photo_file_ids(messages),
            format_application(
                f"Сдача #{rental.rental_no or rental.id}",
                rental.employee_name or employee_name(text_message),
                rental.username or username(text_message),
                rental.user_id,
                model,
                returned.plate,
                rental.created_at,
                rental.planned_return_at,
                rental.return_text,
                days,
                returned.comment,
                rental.total,
            ),
        )
        if rental.total and rental.total > 0:
            notes.append("После списания аренды отправьте /paid ответом на сообщение возврата.")
        summary = await text_message.reply("\n".join(notes))
        storage.update_return_message_id(rental.id, summary.message_id)


async def flush_media_group(media_group_id: str) -> None:
    await asyncio.sleep(config.media_group_wait_seconds)
    buffer = media_groups.pop(media_group_id, None)
    if buffer:
        await process_messages(buffer.messages)


@dp.message(Command("help"))
async def help_command(message: Message) -> None:
    await message.answer(HELP_TEXT)


@dp.message(Command("start"))
async def start_command(message: Message) -> None:
    await message.answer("Выберите действие:", reply_markup=start_menu(is_report_user(message)))


@dp.message(Command("car"))
async def car_command(message: Message) -> None:
    await message.answer("Выберите действие:", reply_markup=start_menu(is_report_user(message)))


@dp.message(Command("status"))
async def status_command(message: Message) -> None:
    await message.answer(f"Активных аренд сейчас: {storage.active_count()}")


@dp.message(Command("chat_id"))
async def chat_id_command(message: Message) -> None:
    await message.answer(f"ID этого чата: {message.chat.id}")


@dp.message(Command("set_work_chat"))
async def set_work_chat_command(message: Message) -> None:
    if not is_report_user(message):
        await message.reply("Рабочий чат может назначить только пользователь с доступом к отчетам.")
        return
    storage.set_setting("work_chat_id", str(message.chat.id))
    await message.answer(f"Готово. Этот чат назначен рабочим для пересылки фото и сводок: {message.chat.id}")


@dp.message(Command("report_month"))
async def report_month_command(message: Message) -> None:
    if not is_report_user(message):
        await message.reply("Отчет может запросить только пользователь с доступом к отчетам.")
        return

    parts = (message.text or "").split()
    year_month = parts[1] if len(parts) > 1 else datetime.now().strftime("%Y-%m")
    rows = storage.monthly_report_rows(year_month)
    data = build_report_xlsx(rows, year_month)
    await message.answer_document(
        BufferedInputFile(data, filename=f"car-rentals-{year_month}.xlsx"),
        caption=f"Отчет за {year_month}: {len(rows)} записей.",
    )


@dp.message(Command("report_employee"))
async def report_employee_command(message: Message) -> None:
    if not is_report_user(message):
        await message.reply("Отчет может запросить только пользователь с доступом к отчетам.")
        return

    parts = (message.text or "").split(maxsplit=2)
    year_month = datetime.now().strftime("%Y-%m")
    query = ""
    if len(parts) >= 2:
        if len(parts[1]) == 7 and parts[1][4] == "-":
            year_month = parts[1]
            query = parts[2] if len(parts) == 3 else ""
        else:
            query = " ".join(parts[1:])
    if not query:
        await message.reply("Укажите сотрудника: `/report_employee 2026-05 @username` или `/report_employee Иван Иванов`.")
        return

    rows = storage.monthly_report_rows(year_month, query)
    data = build_report_xlsx(rows, year_month)
    safe_query = query.strip().lstrip("@").replace(" ", "_")[:40] or "employee"
    await message.answer_document(
        BufferedInputFile(data, filename=f"car-rentals-{year_month}-{safe_query}.xlsx"),
        caption=f"Отчет по сотруднику за {year_month}: {len(rows)} записей.",
    )


@dp.message(Command("employees"))
async def employees_command(message: Message) -> None:
    if not is_report_user(message):
        await message.reply("Список может запросить только пользователь с доступом к отчетам.")
        return

    parts = (message.text or "").split()
    year_month = parts[1] if len(parts) > 1 else datetime.now().strftime("%Y-%m")
    rows = storage.employees_for_month(year_month)
    if not rows:
        await message.reply(f"За {year_month} сотрудников с арендой не нашел.")
        return

    lines = [f"Сотрудники с арендой за {year_month}:"]
    for row in rows:
        username_text = f"@{row['username']}" if row["username"] else f"id {row['user_id']}"
        lines.append(
            f"{row['employee_name'] or 'Без ФИО'} ({username_text}) - "
            f"{row['rentals_count']} аренд, {row['total_days']} дн."
        )
    await message.answer("\n".join(lines))


@dp.callback_query(F.data == "report:month")
async def report_month_button(callback: CallbackQuery) -> None:
    if not is_report_callback(callback):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    year_month = datetime.now().strftime("%Y-%m")
    rows = storage.monthly_report_rows(year_month)
    data = build_report_xlsx(rows, year_month)
    await callback.message.answer_document(
        BufferedInputFile(data, filename=f"car-rentals-{year_month}.xlsx"),
        caption=f"Отчет за {year_month}: {len(rows)} записей.",
    )
    await callback.answer()


@dp.callback_query(F.data == "report:employees")
async def employees_button(callback: CallbackQuery) -> None:
    if not is_report_callback(callback):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    year_month = datetime.now().strftime("%Y-%m")
    rows = storage.employees_for_month(year_month)
    if not rows:
        await callback.message.answer(f"За {year_month} сотрудников с арендой не нашел.")
        await callback.answer()
        return
    lines = [f"Сотрудники с арендой за {year_month}:"]
    for row in rows:
        username_text = f"@{row['username']}" if row["username"] else f"id {row['user_id']}"
        lines.append(
            f"{row['employee_name'] or 'Без ФИО'} ({username_text}) - "
            f"{row['rentals_count']} аренд, {row['total_days']} дн."
        )
    await callback.message.answer("\n".join(lines))
    await callback.answer()


@dp.callback_query(F.data == "report:employee_help")
async def report_employee_help_button(callback: CallbackQuery) -> None:
    if not is_report_callback(callback):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await callback.message.answer(
        "Для выгрузки по сотруднику отправьте команду:\n"
        "`/report_employee 2026-05 @username`\n"
        "или\n"
        "`/report_employee Иван Иванов`"
    )
    await callback.answer()


def build_report_xlsx(rows, year_month: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Аренды"
    headers = [
        "ФИО",
        "Марка модель машины",
        "Гос номер",
        "Дата взятия",
        "Дата сдачи",
        "Кол-во дней",
        "Сумма аренды",
        "Комментарий",
    ]
    sheet.append([f"Отчет по арендам за {year_month}"])
    sheet.append(headers)
    for row in rows:
        _, calculated_total = rental_price(
            row["model"],
            row["days"],
            row["planned_return_at"],
            row["created_at"],
        )
        sheet.append(
            [
                row["employee_name"] or "",
                row["model"],
                row["plate"],
                format_dt(row["created_at"]),
                format_dt(row["returned_at"]),
                report_days(row["days"]),
                row["total"] or calculated_total,
                report_comment(row["take_comment"], row["return_comment"]),
            ]
        )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{sheet.max_row}"
    widths = [28, 24, 16, 20, 20, 12, 16, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@dp.message(Command("paid"))
async def paid_command(message: Message) -> None:
    if not is_approver(message):
        await message.reply("Отметить списание может только согласующий или администратор.")
        return
    if not message.reply_to_message:
        await message.reply("Команду /paid нужно отправить ответом на сообщение о возврате.")
        return
    ok = storage.mark_paid_by_return_message(message.chat.id, message.reply_to_message.message_id)
    await message.reply("Списание отмечено." if ok else "Не нашел возврат по сообщению, на которое вы ответили.")


@dp.callback_query(F.data == "car:take")
async def start_take(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not has_employee_name(callback.from_user):
        await state.update_data(pending_action="take")
        await state.set_state(ProfileFlow.name)
        await callback.message.answer("Введите ваше ФИО один раз, чтобы бот корректно формировал отчеты.")
        await callback.answer()
        return
    await state.set_state(TakeFlow.model)
    await callback.message.answer(f"Выберите марку авто:\n\n{model_legend()}", reply_markup=model_menu())
    await callback.answer()


@dp.message(ProfileFlow.name)
async def save_profile_name(message: Message, state: FSMContext) -> None:
    full_name = " ".join((message.text or "").split())
    if len(full_name.split()) < 2:
        await message.answer("Пожалуйста, укажите ФИО полностью, например: Иванов Иван Иванович.")
        return
    storage.set_employee_name(message.from_user.id, full_name)
    data = await state.get_data()
    if data.get("pending_action") == "take":
        await state.set_state(TakeFlow.model)
        await message.answer("Спасибо, ФИО сохранено.\nВыберите марку авто:", reply_markup=model_menu())
        return
    if data.get("pending_action") == "test_drive":
        await state.set_state(TestDriveFlow.model)
        await message.answer("Спасибо, ФИО сохранено.\nВыберите марку авто для тест драйва:", reply_markup=model_menu("test_model"))
        return
    await state.clear()
    await message.answer("Спасибо, ФИО сохранено. Нажмите /start для продолжения.")


@dp.callback_query(TakeFlow.model, F.data.startswith("take_model:"))
async def take_model_button(callback: CallbackQuery, state: FSMContext) -> None:
    raw_index = callback.data.split(":", 1)[1]
    model = model_by_index(int(raw_index)) if raw_index.isdigit() else None
    if model is None:
        await callback.answer("Не нашел модель", show_alert=True)
        return
    await state.update_data(model=model)
    await state.set_state(TakeFlow.plate)
    await callback.message.answer(f"Выбрано: {model}\nВведите госномер, например: Н537РА126")
    await callback.answer()


@dp.callback_query(F.data.startswith("noop:"))
async def noop_callback(callback: CallbackQuery) -> None:
    await callback.answer()


@dp.message(TakeFlow.model)
async def take_model(message: Message, state: FSMContext) -> None:
    await message.answer("Выберите марку авто кнопкой из списка выше.")


@dp.message(TakeFlow.plate)
async def take_plate(message: Message, state: FSMContext) -> None:
    plate = (message.text or "").strip().upper().replace(" ", "")
    await state.update_data(plate=plate)
    await state.set_state(TakeFlow.return_text)
    await message.answer("Когда возврат? Например: завтра 18:30, 29 августа 20:00, 29.08.2026 12:00.")


@dp.message(TakeFlow.return_text)
async def take_return_text(message: Message, state: FSMContext) -> None:
    plan = parse_return_plan(message.text or "")
    if plan is None:
        await message.answer("Не понял дату возврата. Напишите, например: завтра 18:30, 29 августа 20:00 или 29.08.2026 12:00.")
        return
    await state.update_data(
        return_text=plan.text,
        planned_return_at=plan.planned_at.isoformat(timespec="minutes"),
        days=plan.days,
    )
    await state.set_state(TakeFlow.comment)
    await message.answer(
        f"Посчитал срок аренды: {plan.days} дн.\n"
        "Комментарий при взятии необязателен. Если согласовали, напишите с кем. Можно также указать цель, ночную смену или пропустить.",
        reply_markup=skip_take_comment_menu(),
    )


@dp.message(TakeFlow.comment)
async def take_comment(message: Message, state: FSMContext) -> None:
    await state.update_data(
        take_comment=(message.text or "").strip(),
        photo_count=0,
        first_photo_message_id=None,
        photo_message_ids=[],
        photo_file_ids=[],
    )
    await state.set_state(TakeFlow.photos)
    await message.answer(
        "Пришлите 5 фото: 4 стороны авто и 1 фото приборной панели, где виден уровень топлива."
    )


@dp.callback_query(TakeFlow.comment, F.data == "take:skip_comment")
async def skip_take_comment(callback: CallbackQuery, state: FSMContext) -> None:
    await state.update_data(take_comment="", photo_count=0, first_photo_message_id=None, photo_message_ids=[], photo_file_ids=[])
    await state.set_state(TakeFlow.photos)
    await callback.message.answer(
        "Ок, без комментария. Пришлите 5 фото: 4 стороны авто и 1 фото приборной панели, где виден уровень топлива."
    )
    await callback.answer()


@dp.message(TakeFlow.photos, F.photo)
async def take_photos(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    count = int(data.get("photo_count", 0)) + 1
    first_photo_message_id = data.get("first_photo_message_id") or message.message_id
    photo_message_ids = list(data.get("photo_message_ids", []))
    photo_message_ids.append(message.message_id)
    file_ids = list(data.get("photo_file_ids", []))
    file_ids.append(message.photo[-1].file_id)
    await state.update_data(
        photo_count=count,
        first_photo_message_id=first_photo_message_id,
        photo_message_ids=photo_message_ids,
        photo_file_ids=file_ids,
    )

    if count < 5:
        await message.answer(f"Фото принято: {count}/5.")
        return

    data = await state.get_data()
    fake_text = (
        f"Взял {data['model']} {data['plate']} на {data['days']} дня, "
        f"{data.get('take_comment', '')}, возврат {data['return_text']}"
    )
    take = parse_take(fake_text)
    if take is None:
        await message.answer("Не смог собрать заявку. Начните заново через /car.")
        await state.clear()
        return

    check = check_take_rules(config, take, username(message))
    rate, total = rental_price(take.model, take.days, data.get("planned_return_at"))
    rental_id = storage.create_take(
        chat_id=message.chat.id,
        message_id=first_photo_message_id,
        user_id=message.from_user.id,
        username=username(message),
        employee_name=employee_name(message),
        model=take.model,
        plate=take.plate,
        days=take.days,
        return_text=take.return_text,
        planned_return_at=data.get("planned_return_at"),
        night_shift=take.is_night_shift,
        photo_count=count,
        take_comment=data.get("take_comment"),
        rate=rate,
        total=total,
    )
    rental = storage.get_by_id(rental_id)
    public_no = rental.rental_no if rental and rental.rental_no else rental_id
    notes = [f"Заявка на выдачу #{public_no} зафиксирована: {take.model} {take.plate}, {take.days} дн., сумма {total} руб."]
    if check.warnings:
        notes.extend(check.warnings)
    notes.append("Фото отправлены в рабочий чат." if effective_work_chat_id() else "Рабочий чат не задан, фото никуда не отправлены.")
    await send_to_work_chat(
        message.bot,
        list(data.get("photo_file_ids", [])),
        format_application(
            f"Выдача #{public_no}",
            employee_name(message),
            username(message),
            message.from_user.id,
            take.model,
            take.plate,
            rental.created_at if rental else None,
            rental.planned_return_at if rental else data.get("planned_return_at"),
            take.return_text,
            take.days,
            data.get("take_comment"),
            total,
        ),
    )
    summary = await message.answer("\n".join(notes))
    storage.update_take_message_id(rental_id, summary.message_id)
    await state.clear()


@dp.message(TakeFlow.photos)
async def take_photos_wrong(message: Message) -> None:
    await message.answer("На этом шаге нужны именно фото: 4 стороны авто и приборная панель с топливом.")


@dp.callback_query(F.data == "car:return")
async def start_return(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    rentals = storage.active_for_user(callback.message.chat.id, callback.from_user.id)
    if len(rentals) == 1:
        await state.update_data(plate=rentals[0].plate)
        await state.set_state(ReturnFlow.condition)
        await callback.message.answer(
            f"Сдаем {rentals[0].model} {rentals[0].plate}. Опишите недостатки машины или нажмите кнопку:",
            reply_markup=condition_menu(),
        )
    else:
        await state.set_state(ReturnFlow.plate)
        await callback.message.answer("Введите госномер машины, которую сдаете.")
    await callback.answer()


@dp.callback_query(F.data == "car:test_drive")
async def start_test_drive(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if not has_employee_name(callback.from_user):
        await state.update_data(pending_action="test_drive")
        await state.set_state(ProfileFlow.name)
        await callback.message.answer("Введите ваше ФИО один раз, чтобы бот корректно формировал отчеты.")
        await callback.answer()
        return
    await state.set_state(TestDriveFlow.model)
    await callback.message.answer(f"Выберите марку авто для тест драйва:\n\n{model_legend()}", reply_markup=model_menu("test_model"))
    await callback.answer()


@dp.callback_query(TestDriveFlow.model, F.data.startswith("test_model:"))
async def test_drive_model_button(callback: CallbackQuery, state: FSMContext) -> None:
    raw_index = callback.data.split(":", 1)[1]
    model = model_by_index(int(raw_index)) if raw_index.isdigit() else None
    if model is None:
        await callback.answer("Не нашел модель", show_alert=True)
        return
    await state.update_data(model=model)
    await state.set_state(TestDriveFlow.plate)
    await callback.message.answer(f"Выбрано: {model}\nВведите госномер, например: Н537РА126")
    await callback.answer()


@dp.message(TestDriveFlow.model)
async def test_drive_model(message: Message, state: FSMContext) -> None:
    await message.answer("Выберите марку авто кнопкой из списка выше.")


@dp.message(TestDriveFlow.plate)
async def test_drive_plate(message: Message, state: FSMContext) -> None:
    await state.update_data(plate=(message.text or "").strip().upper().replace(" ", ""))
    await state.set_state(TestDriveFlow.comment)
    await message.answer("Напишите, с кем согласован тест драйв. Это обязательный комментарий.")


@dp.message(TestDriveFlow.comment)
async def test_drive_comment(message: Message, state: FSMContext) -> None:
    comment = (message.text or "").strip()
    if not comment:
        await message.answer("Для тест драйва обязательно нужно указать, с кем он согласован.")
        return
    await state.update_data(
        test_comment=comment,
        photo_count=0,
        first_photo_message_id=None,
        photo_message_ids=[],
        photo_file_ids=[],
    )
    await state.set_state(TestDriveFlow.photos)
    await message.answer("Комментарий принят. Пришлите 5 фото: 4 стороны авто и приборная панель с уровнем топлива.")


@dp.message(TestDriveFlow.photos, F.photo)
async def test_drive_photos(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    count = int(data.get("photo_count", 0)) + 1
    first_photo_message_id = data.get("first_photo_message_id") or message.message_id
    photo_message_ids = list(data.get("photo_message_ids", []))
    photo_message_ids.append(message.message_id)
    file_ids = list(data.get("photo_file_ids", []))
    file_ids.append(message.photo[-1].file_id)
    await state.update_data(
        photo_count=count,
        first_photo_message_id=first_photo_message_id,
        photo_message_ids=photo_message_ids,
        photo_file_ids=file_ids,
    )
    if count < 5:
        await message.answer(f"Фото принято: {count}/5.")
        return

    data = await state.get_data()
    rental_id = storage.create_test_drive(
        chat_id=message.chat.id,
        message_id=first_photo_message_id,
        user_id=message.from_user.id,
        username=username(message),
        employee_name=employee_name(message),
        model=data["model"],
        plate=data["plate"],
        photo_count=count,
        comment=data["test_comment"],
    )
    rental = storage.get_by_id(rental_id)
    public_no = rental.rental_no if rental and rental.rental_no else rental_id
    notes = [f"Тест драйв #{public_no} зафиксирован: {data['model']} {data['plate']}, бесплатно."]
    notes.append("Фото отправлены в рабочий чат." if effective_work_chat_id() else "Рабочий чат не задан, фото никуда не отправлены.")
    await send_to_work_chat(
        message.bot,
        list(data.get("photo_file_ids", [])),
        format_application(
            f"Тест драйв #{public_no}",
            employee_name(message),
            username(message),
            message.from_user.id,
            data["model"],
            data["plate"],
            rental.created_at if rental else None,
            None,
            "тест драйв",
            0,
            f"Согласовано: {data['test_comment']}",
            0,
        ),
    )
    await message.answer("\n".join(notes))
    await state.clear()


@dp.message(TestDriveFlow.photos)
async def test_drive_photos_wrong(message: Message) -> None:
    await message.answer("На этом шаге нужны 5 фото: 4 стороны авто и приборная панель с топливом.")


@dp.message(ReturnFlow.plate)
async def return_plate(message: Message, state: FSMContext) -> None:
    await state.update_data(plate=(message.text or "").strip().upper().replace(" ", ""))
    await state.set_state(ReturnFlow.condition)
    await message.answer("Опишите недостатки машины или нажмите кнопку:", reply_markup=condition_menu())


@dp.callback_query(ReturnFlow.condition, F.data == "return:ok")
async def return_condition_ok(callback: CallbackQuery, state: FSMContext) -> None:
    await state.update_data(
        return_comment="Машина полностью исправна",
        condition_status="ok",
        photo_count=0,
        photo_message_ids=[],
        photo_file_ids=[],
    )
    await state.set_state(ReturnFlow.photos)
    await callback.message.answer("Пришлите 5 фото: 4 стороны авто и приборная панель с уровнем топлива.")
    await callback.answer()


@dp.message(ReturnFlow.condition)
async def return_condition_comment(message: Message, state: FSMContext) -> None:
    comment = (message.text or "").strip()
    if not comment:
        await message.answer("Нужен комментарий о недостатках или кнопка `Машина полностью исправна`.")
        return
    await state.update_data(return_comment=comment, condition_status="comment", photo_count=0, photo_message_ids=[], photo_file_ids=[])
    await state.set_state(ReturnFlow.photos)
    await message.answer("Комментарий принят. Теперь пришлите 5 фото: 4 стороны авто и приборная панель с уровнем топлива.")


@dp.message(ReturnFlow.photos, F.photo)
async def return_photos(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    count = int(data.get("photo_count", 0)) + 1
    first_photo_message_id = data.get("first_photo_message_id") or message.message_id
    photo_message_ids = list(data.get("photo_message_ids", []))
    photo_message_ids.append(message.message_id)
    file_ids = list(data.get("photo_file_ids", []))
    file_ids.append(message.photo[-1].file_id)
    await state.update_data(
        photo_count=count,
        first_photo_message_id=first_photo_message_id,
        photo_message_ids=photo_message_ids,
        photo_file_ids=file_ids,
    )
    if count < 5:
        await message.answer(f"Фото принято: {count}/5.")
        return

    data = await state.get_data()
    rental = storage.mark_returned(
        message.chat.id,
        first_photo_message_id,
        data["plate"],
        count,
        return_comment=data.get("return_comment"),
        condition_status=data.get("condition_status"),
    )
    if rental is None:
        await message.answer("Не нашел активную выдачу по этому номеру. Проверьте номер или обратитесь к ответственному.")
        await state.clear()
        return
    notes = [f"Возврат зафиксирован: {rental.model} {rental.plate}, сумма {rental.total or 0} руб."]
    notes.append("Фото отправлены в рабочий чат." if effective_work_chat_id() else "Рабочий чат не задан, фото никуда не отправлены.")
    await send_to_work_chat(
        message.bot,
        list(data.get("photo_file_ids", [])),
        format_application(
            f"Сдача #{rental.rental_no or rental.id}",
            rental.employee_name or employee_name(message),
            rental.username or username(message),
            rental.user_id,
            rental.model,
            rental.plate,
            rental.created_at,
            rental.planned_return_at,
            rental.return_text,
            rental.days,
            data.get("return_comment"),
            rental.total,
        ),
    )
    if rental.total and rental.total > 0:
        notes.append("После списания аренды ответственный отправляет /paid ответом на это сообщение.")
    summary = await message.answer("\n".join(notes))
    storage.update_return_message_id(rental.id, summary.message_id)
    await state.clear()


@dp.message(ReturnFlow.photos)
async def return_photos_wrong(message: Message) -> None:
    await message.answer("На этом шаге нужны 5 фото: 4 стороны авто и приборная панель с топливом.")


@dp.message(F.reply_to_message)
async def approval_message(message: Message) -> None:
    approval = parse_approval(message.text or message.caption or "")
    if approval is None:
        await collect_or_process(message)
        return
    if not is_approver(message):
        await message.reply("Согласование может отправить только Сергей, администратор или @stax_ru.")
        return
    rental = storage.get_by_take_message(message.chat.id, message.reply_to_message.message_id)
    if rental is None:
        await message.reply("Не нашел выдачу в сообщении, на которое вы ответили.")
        return
    storage.set_approval(
        chat_id=message.chat.id,
        take_message_id=message.reply_to_message.message_id,
        approval_message_id=message.message_id,
        rate=approval.rate,
        total=approval.total,
        approval_status=approval.status,
    )
    take_from_reply = parse_take(message.reply_to_message.caption or message.reply_to_message.text or "")
    warnings = []
    if take_from_reply is not None:
        expected = check_take_rules(config, take_from_reply, rental.username)
        warnings = validate_approval(expected, approval)
    answer = ["Согласование зафиксировано."]
    answer.extend(warnings)
    await message.reply("\n".join(answer))


@dp.message()
async def collect_or_process(message: Message) -> None:
    if message.media_group_id:
        buffer = media_groups[message.media_group_id]
        buffer.messages.append(message)
        if buffer.task is None:
            buffer.task = asyncio.create_task(flush_media_group(message.media_group_id))
        return
    await process_messages([message])


async def main() -> None:
    bot = Bot(token=config.bot_token)
    await bot.set_my_commands(
        [
            BotCommand(command="start", description="Открыть меню"),
            BotCommand(command="help", description="Показать инструкцию"),
        ]
    )
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
